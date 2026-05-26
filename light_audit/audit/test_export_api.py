"""Tests for US-050: Audit spreadsheet export (CSV/XLSX)."""
import csv
import io
from http import HTTPStatus

import openpyxl
import pytest
from django.contrib.auth import get_user_model

from light_audit.audit.models import AuditFlag
from light_audit.audit.models import AuditVersion
from light_audit.audit.models import Building
from light_audit.audit.models import Floor
from light_audit.audit.models import LogEntry
from light_audit.audit.models import Project
from light_audit.audit.models import Room

User = get_user_model()


@pytest.fixture
def user(db):
    return User.objects.create_user(email="export@example.com", password="pass")  # noqa: S106


@pytest.fixture
def project(db, user):
    return Project.objects.create(name="Export Project", owner=user)


@pytest.fixture
def building(db, project):
    return Building.objects.create(name="Export Building", project=project)


@pytest.fixture
def version(db, building, user):
    return AuditVersion.objects.create(building=building, created_by=user)


@pytest.fixture
def floor(db, building, version):
    return Floor.objects.create(
        building=building, audit_version=version, name="Floor 1", level=1,
    )


@pytest.fixture
def room(db, floor, version):
    return Room.objects.create(
        floor=floor, audit_version=version, name="Room A",
    )


@pytest.fixture
def log_entry(db, room, version):
    return LogEntry.objects.create(
        room=room,
        audit_version=version,
        fixture_id="E1",
        qty=4,
        wattage="25.50",
        description="LED Troffer",
        flag_embb=True,
    )


# --- Auth ---

@pytest.mark.django_db
def test_export_xlsx_requires_auth(client, version):
    response = client.get(f"/api/audit-versions/{version.pk}/export/xlsx/")
    assert response.status_code == HTTPStatus.UNAUTHORIZED


@pytest.mark.django_db
def test_export_csv_requires_auth(client, version):
    response = client.get(f"/api/audit-versions/{version.pk}/export/csv/")
    assert response.status_code == HTTPStatus.UNAUTHORIZED


# --- XLSX ---

@pytest.mark.django_db
def test_export_xlsx_returns_file(client, user, version, log_entry):
    client.force_login(user)
    response = client.get(f"/api/audit-versions/{version.pk}/export/xlsx/")
    assert response.status_code == HTTPStatus.OK
    content_type = response.headers.get("Content-Type", "")
    assert "spreadsheetml" in content_type
    assert "xlsx" in response.headers.get("Content-Disposition", "")


@pytest.mark.django_db
def test_export_xlsx_column_shape(client, user, version, log_entry):
    client.force_login(user)
    response = client.get(f"/api/audit-versions/{version.pk}/export/xlsx/")
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    assert "fixture_id" in header
    assert "active_flags" in header
    assert "floor" in header
    assert "room" in header
    # one data row
    assert len(rows) == 2  # noqa: PLR2004


@pytest.mark.django_db
def test_export_xlsx_data_values(client, user, version, log_entry):
    client.force_login(user)
    response = client.get(f"/api/audit-versions/{version.pk}/export/xlsx/")
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    data = dict(zip(header, rows[1], strict=False))
    assert data["fixture_id"] == "E1"
    assert data["qty"] == 4  # noqa: PLR2004
    assert data["wattage"] == 25.5  # noqa: PLR2004
    assert data["flag_embb"] is True


@pytest.mark.django_db
def test_export_xlsx_active_flags_column(client, user, version, log_entry):
    AuditFlag.objects.create(
        log_entry=log_entry,
        audit_version=version,
        severity="warn",
        message="Check ballast",
        status="active",
    )
    client.force_login(user)
    response = client.get(f"/api/audit-versions/{version.pk}/export/xlsx/")
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    data = dict(zip(header, rows[1], strict=False))
    assert "WARN: Check ballast" in data["active_flags"]


@pytest.mark.django_db
def test_export_xlsx_dismissed_flags_excluded(client, user, version, log_entry):
    AuditFlag.objects.create(
        log_entry=log_entry,
        audit_version=version,
        severity="info",
        message="Dismissed note",
        status="dismissed",
    )
    client.force_login(user)
    response = client.get(f"/api/audit-versions/{version.pk}/export/xlsx/")
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    data = dict(zip(header, rows[1], strict=False))
    assert data["active_flags"] in ("", None)  # openpyxl returns None for empty cells


# --- CSV ---

@pytest.mark.django_db
def test_export_csv_returns_file(client, user, version, log_entry):
    client.force_login(user)
    response = client.get(f"/api/audit-versions/{version.pk}/export/csv/")
    assert response.status_code == HTTPStatus.OK
    assert "text/csv" in response.headers.get("Content-Type", "")
    assert "csv" in response.headers.get("Content-Disposition", "")


@pytest.mark.django_db
def test_export_csv_column_shape(client, user, version, log_entry):
    client.force_login(user)
    response = client.get(f"/api/audit-versions/{version.pk}/export/csv/")
    text = response.content.decode()
    reader = list(csv.reader(io.StringIO(text)))
    header = reader[0]
    assert "fixture_id" in header
    assert "active_flags" in header
    row_count = len(reader)
    expected_rows = 2  # header + 1 entry
    assert row_count == expected_rows


@pytest.mark.django_db
def test_export_csv_data_values(client, user, version, log_entry):
    client.force_login(user)
    response = client.get(f"/api/audit-versions/{version.pk}/export/csv/")
    text = response.content.decode()
    reader = list(csv.reader(io.StringIO(text)))
    header = reader[0]
    data = dict(zip(header, reader[1], strict=False))
    assert data["fixture_id"] == "E1"
    assert data["qty"] == "4"
    assert data["flag_embb"] == "True"


@pytest.mark.django_db
def test_export_404_on_unknown_version(client, user):
    client.force_login(user)
    response = client.get("/api/audit-versions/99999/export/xlsx/")
    assert response.status_code == HTTPStatus.NOT_FOUND
